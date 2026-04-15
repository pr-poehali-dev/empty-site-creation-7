"""Экспорт заявки в Excel-файл"""
import json
import os
import base64
import io
import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def get_db():
    return psycopg2.connect(os.environ['DATABASE_URL'])


def get_user_by_token(cur, token):
    cur.execute(
        """SELECT u.id, u.phone, u.role FROM users u
           JOIN user_sessions s ON s.user_id = u.id
           WHERE s.token = %s AND s.expires_at > NOW()""",
        (token,)
    )
    return cur.fetchone()


def handler(event: dict, context) -> dict:
    """Экспорт оптовой заявки в Excel-файл с формулами"""
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type, Authorization, X-Authorization',
                'Access-Control-Max-Age': '86400'
            },
            'body': ''
        }

    headers = {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'}
    params = event.get('queryStringParameters') or {}
    order_id = params.get('id')
    if not order_id:
        return {'statusCode': 400, 'headers': headers, 'body': json.dumps({'error': 'Не указан id заявки'})}

    req_headers = event.get('headers', {})
    auth = req_headers.get('X-Authorization', '') or req_headers.get('Authorization', '')
    token = auth.replace('Bearer ', '').strip()

    conn = get_db()
    cur = conn.cursor()

    user = get_user_by_token(cur, token)
    if not user:
        cur.close(); conn.close()
        return {'statusCode': 401, 'headers': headers, 'body': json.dumps({'error': 'Не авторизован'})}

    cur.execute(
        """SELECT o.id, o.customer_name, o.comment, o.status, o.total_amount, o.created_at
           FROM wholesale_orders o WHERE o.id = %s""",
        (order_id,)
    )
    order = cur.fetchone()
    if not order:
        cur.close(); conn.close()
        return {'statusCode': 404, 'headers': headers, 'body': json.dumps({'error': 'Заявка не найдена'})}

    cur.execute(
        """SELECT oi.product_id, COALESCE(oi.item_name, p.name), p.article, oi.quantity, oi.price, oi.temp_product_id,
                  (SELECT pb.barcode FROM product_barcodes pb WHERE pb.product_id = oi.product_id LIMIT 1)
           FROM wholesale_order_items oi
           JOIN products p ON p.id = oi.product_id
           WHERE oi.order_id = %s
           ORDER BY oi.id""",
        (order_id,)
    )
    items = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Заявка {order_id}"

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    title_font = Font(bold=True, size=13)

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    ws.merge_cells('A1:B1')
    ws['A1'] = f"Оптовик: {order[1]}"
    ws['A1'].font = title_font

    ws.merge_cells('D1:F1')
    ws['D1'] = f"Комментарий: {order[2] or '—'}"
    ws['D1'].font = Font(size=11)
    ws['D1'].alignment = Alignment(horizontal='right')

    ws.merge_cells('A2:B2')
    ws['A2'] = f"Заявка №{order[0]} от {str(order[5])[:10]}"
    ws['A2'].font = Font(size=10, color='666666')

    table_headers = ['Штрихкод', 'Наименование', 'Артикул', 'Кол-во', 'Цена', 'Сумма']
    start_row = 4
    for col_idx, h in enumerate(table_headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center')

    for i, item in enumerate(items):
        row = start_row + 1 + i
        is_temp = item[5] is not None or item[0] == 19
        barcode = item[6] if not is_temp else ''
        article = item[2] if item[2] != '__TEMP__' else ''

        ws.cell(row=row, column=1, value=barcode or '').border = border
        ws.cell(row=row, column=2, value=item[1]).border = border
        ws.cell(row=row, column=3, value=article or '').border = border
        qty_cell = ws.cell(row=row, column=4, value=item[3])
        qty_cell.border = border
        qty_cell.alignment = Alignment(horizontal='center')
        price_cell = ws.cell(row=row, column=5, value=float(item[4]))
        price_cell.border = border
        price_cell.number_format = '#,##0.00'
        sum_cell = ws.cell(row=row, column=6)
        sum_cell.value = f'=D{row}*E{row}'
        sum_cell.border = border
        sum_cell.number_format = '#,##0.00'

    total_row = start_row + 1 + len(items)
    ws.merge_cells(f'A{total_row}:E{total_row}')
    total_label = ws.cell(row=total_row, column=1, value='ИТОГО')
    total_label.font = Font(bold=True, size=12)
    total_label.alignment = Alignment(horizontal='right')
    total_label.border = border
    for col in range(2, 6):
        ws.cell(row=total_row, column=col).border = border

    first_data = start_row + 1
    last_data = start_row + len(items)
    total_cell = ws.cell(row=total_row, column=6)
    total_cell.value = f'=SUM(F{first_data}:F{last_data})'
    total_cell.font = Font(bold=True, size=12)
    total_cell.border = border
    total_cell.number_format = '#,##0.00'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    file_b64 = base64.b64encode(buf.read()).decode()

    return {
        'statusCode': 200,
        'headers': headers,
        'body': json.dumps({
            'file': file_b64,
            'filename': f'Заявка_{order_id}_{order[1]}.xlsx'
        })
    }
