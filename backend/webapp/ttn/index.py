import json
import os
import boto3
import openpyxl
from io import BytesIO


def handler(event: dict, context) -> dict:
    '''Разбирает загруженный Excel-файл ТТН: ячейки, текст, объединённые диапазоны.'''
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
                'Access-Control-Allow-Headers': 'Content-Type',
                'Access-Control-Max-Age': '86400',
            },
            'body': '',
        }

    params = event.get('queryStringParameters') or {}
    s3_key = params.get('s3_key', 'ttn/cea0379d6d7e4c9aae4f0ca404e0bbc3.xlsx')

    s3 = boto3.client(
        's3',
        endpoint_url='https://bucket.poehali.dev',
        aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
        aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    )
    obj = s3.get_object(Bucket='files', Key=s3_key)
    data = obj['Body'].read()

    wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    ws = wb.active

    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != '':
                cells.append({'coord': cell.coordinate, 'value': str(cell.value)})

    merged = [str(r) for r in ws.merged_cells.ranges]

    result = {
        'sheet': ws.title,
        'dimensions': ws.dimensions,
        'max_row': ws.max_row,
        'max_col': ws.max_column,
        'cells': cells,
        'merged_ranges': merged,
    }

    print('TTN_PARSE_START')
    print(json.dumps(result, ensure_ascii=False))
    print('TTN_PARSE_END')

    return {
        'statusCode': 200,
        'headers': {'Access-Control-Allow-Origin': '*', 'Content-Type': 'application/json'},
        'isBase64Encoded': False,
        'body': json.dumps(result, ensure_ascii=False),
    }